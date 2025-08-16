import streamlit as st
import pandas as pd
import re
import os
from datetime import datetime
import io
import plotly.express as px
import plotly.graph_objects as go
from collections import Counter

class ReservationMatcherWeb:
    def __init__(self):
        self.meituan_file = None
        self.reservation_file = None
        self.merged_df = pd.DataFrame()
        self.original_df = pd.DataFrame()
        
    def smart_table_match(self, reservation_table, meituan_table):
        """æ™ºèƒ½æ¡Œç‰Œå·åŒ¹é…å‡½æ•°"""
        # æå–æ•°å­—éƒ¨åˆ†
        def extract_numbers(table_str):
            if pd.isna(table_str):
                return None
            numbers = re.findall(r'\d+', str(table_str))
            return ''.join(numbers) if numbers else None
        
        # åˆ¤æ–­æ˜¯å¦ä¸ºå¤–å–è®¢å•
        def is_takeout(table_str):
            if pd.isna(table_str):
                return False
            table_str = str(table_str).lower()
            takeout_keywords = ['å¤–å–', 'takeout', 'é…é€', 'æ‰“åŒ…']
            return any(keyword in table_str for keyword in takeout_keywords)
        
        # å®Œå…¨åŒ¹é…ï¼ˆæœ€é«˜ä¼˜å…ˆçº§ï¼‰
        if str(reservation_table) == str(meituan_table):
            return True, "å®Œå…¨åŒ¹é…"
        
        # æ•°å­—éƒ¨åˆ†åŒ¹é…ï¼ˆåŒ…æ‹¬å¤–å–è®¢å•ï¼‰
        res_numbers = extract_numbers(reservation_table)
        mt_numbers = extract_numbers(meituan_table)
        
        if res_numbers and mt_numbers and res_numbers == mt_numbers:
            # åŒºåˆ†å¤–å–å’Œå ‚é£Ÿçš„æ•°å­—åŒ¹é…
            if is_takeout(meituan_table):
                return True, "å¤–å–åŒ¹é…"
            else:
                return True, "æ•°å­—åŒ¹é…"
        
        return False, "æ— åŒ¹é…"
    
    def show_record_details(self, selected_record, display_df, selected_idx):
        """æ˜¾ç¤ºé€‰ä¸­è®°å½•çš„è¯¦ç»†ä¿¡æ¯"""
        st.divider()
        st.subheader("ğŸ” è®°å½•è¯¦æƒ…")
        
        # åˆ›å»ºä¸¤åˆ—å¸ƒå±€
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### ğŸ“‹ é¢„è®¢ä¿¡æ¯")
            reservation_info = {
                "æ—¥æœŸ": selected_record.get('æ—¥æœŸ', ''),
                "æ¡Œç‰Œå·": selected_record.get('æ¡Œç‰Œå·', ''),
                "é¢„è®¢äºº": selected_record.get('é¢„è®¢äºº', ''),
                "å¸‚åˆ«": selected_record.get('å¸‚åˆ«', ''),
                "åŒ¹é…çŠ¶æ€": selected_record.get('åŒ¹é…çŠ¶æ€', ''),
                "åŒ¹é…ç±»å‹": selected_record.get('åŒ¹é…ç±»å‹', '')
            }
            
            for key, value in reservation_info.items():
                st.text(f"{key}: {value}")
        
        with col2:
            st.markdown("### ğŸ›’ ç¾å›¢è®¢å•ä¿¡æ¯")
            if selected_record.get('åŒ¹é…çŠ¶æ€') == 'å·²åŒ¹é…':
                meituan_info = {
                    "ä¸‹å•æ—¶é—´": selected_record.get('ä¸‹å•æ—¶é—´', ''),
                    "æ¡Œç‰Œå·": selected_record.get('æ¡Œç‰Œå·', ''),
                    "æ”¯ä»˜åˆè®¡": selected_record.get('æ”¯ä»˜åˆè®¡', ''),
                    "ç»“è´¦æ–¹å¼": selected_record.get('ç»“è´¦æ–¹å¼', ''),
                    "ä¸‹å•æ—¶é—´æ ¼å¼åŒ–": selected_record.get('ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–', '')
                }
                
                for key, value in meituan_info.items():
                    st.text(f"{key}: {value}")
                
                # ç§»é™¤åŒ¹é…æŒ‰é’®
                st.markdown("---")
                if st.button("âŒ ç§»é™¤æ­¤åŒ¹é…", key=f"remove_{selected_idx}", type="secondary"):
                    self.remove_match(selected_record, selected_idx)
                    st.rerun()
            else:
                st.info("æ­¤è®°å½•æœªåŒ¹é…åˆ°ç¾å›¢è®¢å•")
    
    def remove_match(self, selected_record, selected_idx):
        """ç§»é™¤åŒ¹é…è®°å½•"""
        try:
            # åœ¨merged_dfä¸­æ‰¾åˆ°å¯¹åº”è®°å½•å¹¶ç§»é™¤åŒ¹é…ä¿¡æ¯
            mask = (
                (self.merged_df['æ—¥æœŸ'] == selected_record['æ—¥æœŸ']) &
                (self.merged_df['æ¡Œç‰Œå·'] == selected_record['æ¡Œç‰Œå·']) &
                (self.merged_df['é¢„è®¢äºº'] == selected_record['é¢„è®¢äºº']) &
                (self.merged_df['å¸‚åˆ«'] == selected_record['å¸‚åˆ«'])
            )
            
            # æ›´æ–°åŒ¹é…çŠ¶æ€å’Œç›¸å…³å­—æ®µ
            self.merged_df.loc[mask, 'åŒ¹é…çŠ¶æ€'] = 'æœªåŒ¹é…'
            self.merged_df.loc[mask, 'åŒ¹é…ç±»å‹'] = 'æœªåŒ¹é…'
            self.merged_df.loc[mask, 'æ”¯ä»˜åˆè®¡'] = None
            self.merged_df.loc[mask, 'ä¸‹å•æ—¶é—´'] = None
            self.merged_df.loc[mask, 'ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–'] = None
            self.merged_df.loc[mask, 'ç»“è´¦æ–¹å¼'] = None
            
            st.success("âœ… å·²æˆåŠŸç§»é™¤åŒ¹é…")
            
        except Exception as e:
            st.error(f"âŒ ç§»é™¤åŒ¹é…å¤±è´¥: {str(e)}")
        
    def load_files(self):
        """æ–‡ä»¶ä¸Šä¼ ç•Œé¢"""
        # ç¾å›¢è®¢å•æ–‡ä»¶ä¸Šä¼ 
        st.write("**ç¾å›¢è®¢å•æ–‡ä»¶**")
        meituan_uploaded = st.file_uploader(
            "é€‰æ‹©ç¾å›¢è®¢å•Excelæ–‡ä»¶", 
            type=['xlsx', 'xls'],
            key="meituan"
        )
        
        if meituan_uploaded:
            try:
                # å°è¯•ä¸åŒçš„headerè®¾ç½®æ¥è¯»å–ç¾å›¢æ–‡ä»¶
                meituan_df = None
                for header_row in [2, 1, 0, None]:
                    try:
                        temp_df = pd.read_excel(meituan_uploaded, header=header_row)
                        # æ£€æŸ¥æ˜¯å¦åŒ…å«å…³é”®åˆ—
                        if any('è¥ä¸šæ—¥æœŸ' in str(col) for col in temp_df.columns) and \
                           any('æ¡Œç‰Œå·' in str(col) for col in temp_df.columns):
                            meituan_df = temp_df
                            break
                    except:
                        continue
                
                if meituan_df is None:
                    st.error("æ— æ³•è¯†åˆ«ç¾å›¢æ–‡ä»¶æ ¼å¼ï¼Œè¯·æ£€æŸ¥æ–‡ä»¶æ˜¯å¦æ­£ç¡®")
                    return
                
                self.meituan_file = meituan_df
                
                # æ¸…ç†æ•°æ®ï¼šç§»é™¤å®Œå…¨ç©ºçš„åˆ—å’Œè¡Œ
                self.meituan_file = self.meituan_file.dropna(how='all', axis=1)  # åˆ é™¤å…¨ç©ºåˆ—
                self.meituan_file = self.meituan_file.dropna(how='all', axis=0)  # åˆ é™¤å…¨ç©ºè¡Œ
                
                # è½¬æ¢æ‰€æœ‰åˆ—ä¸ºå­—ç¬¦ä¸²ç±»å‹ä»¥é¿å…ç±»å‹å†²çª
                for col in self.meituan_file.columns:
                    if self.meituan_file[col].dtype == 'object':
                        self.meituan_file[col] = self.meituan_file[col].astype(str)
                    
                # æ™ºèƒ½æ£€æµ‹åˆ—å
                date_col = None
                table_col = None
                customer_col = None
                
                for col in self.meituan_file.columns:
                    if any(keyword in str(col) for keyword in ['è¥ä¸šæ—¥æœŸ', 'æ—¥æœŸ', 'date']):
                        date_col = col
                    if any(keyword in str(col) for keyword in ['æ¡Œç‰Œå·', 'æ¡Œå·', 'å°å·']):
                        table_col = col
                    if any(keyword in str(col) for keyword in ['å®¢æˆ·', 'å§“å', 'é¡¾å®¢']):
                        customer_col = col
                
                missing_cols = []
                if not date_col: missing_cols.append('æ—¥æœŸç›¸å…³åˆ—')
                if not table_col: missing_cols.append('æ¡Œç‰Œå·ç›¸å…³åˆ—')
                
                if missing_cols:
                    st.error(f"ç¼ºå°‘å¿…è¦åˆ—: {', '.join(missing_cols)}")
                else:
                    st.success(f"âœ… ç¾å›¢æ–‡ä»¶å·²åŠ è½½ ({len(self.meituan_file)} æ¡è®°å½•)")
                    
                    with st.expander("é¢„è§ˆç¾å›¢æ•°æ®", expanded=False):
                        # åˆ›å»ºæ˜¾ç¤ºç”¨çš„DataFrameå‰¯æœ¬
                        display_df = self.meituan_file.copy()
                        
                        # æ·»åŠ æ°´å¹³æ»šåŠ¨æ ·å¼
                        st.markdown("""
                        <style>
                        .stDataFrame {
                            overflow-x: auto;
                        }
                        .stDataFrame > div {
                            overflow-x: auto;
                        }
                        </style>
                        """, unsafe_allow_html=True)
                        
                        st.dataframe(display_df, use_container_width=True)
                    
            except Exception as e:
                st.error(f"ç¾å›¢æ–‡ä»¶åŠ è½½å¤±è´¥: {str(e)}")
            
            st.divider()
            
            # é¢„è®¢è®°å½•æ–‡ä»¶ä¸Šä¼ 
            st.write("**é¢„è®¢è®°å½•æ–‡ä»¶**")
            reservation_uploaded = st.file_uploader(
                "é€‰æ‹©é¢„è®¢è®°å½•Excelæ–‡ä»¶", 
                type=['xlsx', 'xls'],
                key="reservation"
            )
            
            if reservation_uploaded:
                try:
                    # è¯»å–Excelæ–‡ä»¶çš„æ‰€æœ‰å·¥ä½œè¡¨
                    excel_file = pd.ExcelFile(reservation_uploaded)
                    all_sheets_data = []
                    
                    # ç®€åŒ–æ˜¾ç¤ºä¿¡æ¯
                    valid_sheets = 0
                    total_records = 0
                    
                    # é€ä¸ªè¯»å–æ¯ä¸ªå·¥ä½œè¡¨
                    for sheet_name in excel_file.sheet_names:
                        try:
                            sheet_df = pd.read_excel(reservation_uploaded, sheet_name=sheet_name)
                            
                            # æ¸…ç†æ•°æ®ï¼šç§»é™¤å®Œå…¨ç©ºçš„åˆ—å’Œè¡Œ
                            sheet_df = sheet_df.dropna(how='all', axis=1)  # åˆ é™¤å…¨ç©ºåˆ—
                            sheet_df = sheet_df.dropna(how='all', axis=0)  # åˆ é™¤å…¨ç©ºè¡Œ
                            
                            # å¦‚æœå·¥ä½œè¡¨æœ‰æ•°æ®ï¼Œæ·»åŠ åˆ°åˆ—è¡¨ä¸­
                            if not sheet_df.empty:
                                # æ·»åŠ å·¥ä½œè¡¨åç§°åˆ—ç”¨äºæ ‡è¯†æ•°æ®æ¥æº
                                sheet_df['æ•°æ®æ¥æºå·¥ä½œè¡¨'] = sheet_name
                                all_sheets_data.append(sheet_df)
                                valid_sheets += 1
                                total_records += len(sheet_df)
                        except Exception as e:
                            continue  # é™é»˜è·³è¿‡é”™è¯¯çš„å·¥ä½œè¡¨
                    
                    # åˆå¹¶æ‰€æœ‰å·¥ä½œè¡¨çš„æ•°æ®
                    if all_sheets_data:
                        self.reservation_file = pd.concat(all_sheets_data, ignore_index=True)
                        
                        # è½¬æ¢æ‰€æœ‰åˆ—ä¸ºå­—ç¬¦ä¸²ç±»å‹ä»¥é¿å…ç±»å‹å†²çª
                        for col in self.reservation_file.columns:
                            if self.reservation_file[col].dtype == 'object':
                                self.reservation_file[col] = self.reservation_file[col].astype(str)
                        
                        st.success(f"âœ… é¢„è®¢æ–‡ä»¶å·²åŠ è½½ ({len(self.reservation_file)} æ¡è®°å½•)")
                    else:
                        st.error("æ²¡æœ‰æ‰¾åˆ°æœ‰æ•ˆæ•°æ®")
                        self.reservation_file = pd.DataFrame()
                    
                    with st.expander("ğŸ‘€ é¢„è§ˆé¢„è®¢æ•°æ®", expanded=False):
                        # åˆ›å»ºæ˜¾ç¤ºç”¨çš„DataFrameå‰¯æœ¬
                        display_df = self.reservation_file.copy()
                        
                        # æ·»åŠ æ°´å¹³æ»šåŠ¨æ ·å¼
                        st.markdown("""
                        <style>
                        .stDataFrame {
                            overflow-x: auto;
                        }
                        .stDataFrame > div {
                            overflow-x: auto;
                        }
                        </style>
                        """, unsafe_allow_html=True)
                        
                        st.dataframe(display_df, use_container_width=True)
                        
                except Exception as e:
                    st.error(f"âŒ é¢„è®¢æ–‡ä»¶åŠ è½½å¤±è´¥: {str(e)}")
            

    
    def validate_files(self):
        """éªŒè¯æ–‡ä»¶æ˜¯å¦å·²åŠ è½½"""
        if self.meituan_file is None or self.reservation_file is None:
            return False, "è¯·å…ˆä¸Šä¼ ç¾å›¢è®¢å•æ–‡ä»¶å’Œé¢„è®¢è®°å½•æ–‡ä»¶"
        
        if self.meituan_file.empty or self.reservation_file.empty:
            return False, "ä¸Šä¼ çš„æ–‡ä»¶ä¸ºç©ºï¼Œè¯·æ£€æŸ¥æ–‡ä»¶å†…å®¹"
        
        return True, "æ–‡ä»¶éªŒè¯é€šè¿‡"
    
    def match_data(self):
        """æ•°æ®åŒ¹é…æ ¸å¿ƒé€»è¾‘ - ä½¿ç”¨ä¸æ¡Œé¢ç‰ˆå®Œå…¨ç›¸åŒçš„åŒ¹é…ç®—æ³•"""
        try:
            # è¯»å–ç¾å›¢æ•°æ® - ä½¿ç”¨ä¸æ¡Œé¢ç‰ˆç›¸åŒçš„å¤„ç†æ–¹å¼
            df = self.meituan_file.copy()
            
            # æ•°æ®æ¸…æ´—å’Œé¢„å¤„ç†
            df = df[df['è®¢å•çŠ¶æ€'] == 'å·²ç»“è´¦']
            df = df[df['è¥ä¸šæ—¥æœŸ'] != '--']
            
            # æ”¹è¿›çš„æ”¯ä»˜é‡‘é¢æå–
            def extract_payment(payment_str):
                if pd.isna(payment_str):
                    return None
                try:
                    # æŸ¥æ‰¾æ‰€æœ‰æ•°å­—ï¼ˆåŒ…æ‹¬è´Ÿæ•°å’Œå°æ•°ï¼‰
                    numbers = re.findall(r'-?\d+\.?\d*', str(payment_str))
                    if numbers:
                        return float(numbers[0])
                except (ValueError, IndexError):
                    pass
                return None
                
            df['æ”¯ä»˜åˆè®¡'] = df['ç»“è´¦æ–¹å¼'].apply(extract_payment)
            df['è¥ä¸šæ—¥æœŸ'] = pd.to_datetime(df['è¥ä¸šæ—¥æœŸ'], errors='coerce')
            df['ä¸‹å•æ—¶é—´'] = pd.to_datetime(df['ä¸‹å•æ—¶é—´'], errors='coerce')
            
            # æ ¹æ®ä¸‹å•æ—¶é—´åˆ¤æ–­å¸‚åˆ«
            def determine_market_period(order_time):
                if pd.isna(order_time):
                    return None
                try:
                    hour = pd.to_datetime(order_time).hour
                    # åˆå¸‚: 6:00-16:00, æ™šå¸‚: 16:00-24:00
                    if 6 <= hour < 16:
                        return 'åˆå¸‚'
                    elif 16 <= hour <= 23:
                        return 'æ™šå¸‚'
                    else:
                        return None  # éè¥ä¸šæ—¶é—´
                except:
                    return None
            
            df['å¸‚åˆ«'] = df['ä¸‹å•æ—¶é—´'].apply(determine_market_period)
            
            # é€‰æ‹©éœ€è¦çš„åˆ—ï¼Œä¿ç•™ä¸‹å•æ—¶é—´å’Œç»“è´¦æ–¹å¼ç”¨äºæ˜¾ç¤º
            mt_df = df[['è¥ä¸šæ—¥æœŸ', 'æ¡Œç‰Œå·', 'ä¸‹å•æ—¶é—´', 'æ”¯ä»˜åˆè®¡', 'å¸‚åˆ«', 'ç»“è´¦æ–¹å¼']].copy()
            # è¿‡æ»¤æ‰éè¥ä¸šæ—¶é—´çš„è®¢å•
            mt_df = mt_df[mt_df['å¸‚åˆ«'].notna()]
            
            # æ ¼å¼åŒ–ä¸‹å•æ—¶é—´ä¸ºæ›´æ˜“è¯»çš„æ ¼å¼
            try:
                mt_df['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–'] = mt_df['ä¸‹å•æ—¶é—´'].dt.strftime('%H:%M:%S')
            except AttributeError:
                # å¦‚æœä¸æ˜¯datetimeç±»å‹ï¼Œå°è¯•è½¬æ¢åæ ¼å¼åŒ–
                mt_df['ä¸‹å•æ—¶é—´'] = pd.to_datetime(mt_df['ä¸‹å•æ—¶é—´'], errors='coerce')
                mt_df['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–'] = mt_df['ä¸‹å•æ—¶é—´'].dt.strftime('%H:%M:%S')
            
            # æå–ä¸‹å•æ—¶é—´çš„æ—¥æœŸéƒ¨åˆ†ç”¨äºåŒ¹é…
            mt_df['ä¸‹å•æ—¥æœŸ'] = mt_df['ä¸‹å•æ—¶é—´'].dt.date
            
            # è¯»å–é¢„è®¢æ•°æ®
            merged_all = pd.DataFrame()
            
            # å¤„ç†é¢„è®¢æ•°æ® - æ”¯æŒå¤šå·¥ä½œè¡¨
            if hasattr(self.reservation_file, 'sheet_names'):
                # å¦‚æœæ˜¯ExcelFileå¯¹è±¡ï¼Œå¤„ç†å¤šä¸ªå·¥ä½œè¡¨
                for sheet_name in self.reservation_file.sheet_names:
                    try:
                        day_df = pd.read_excel(self.reservation_file, sheet_name=sheet_name)
                        
                        # æ£€æŸ¥å¿…è¦çš„åˆ—æ˜¯å¦å­˜åœ¨
                        required_cols = ['å§“å', 'é¢„è®¢äºº']
                        missing_cols = [col for col in required_cols if col not in day_df.columns]
                        if missing_cols:
                            continue
                            
                        # æ•°æ®æ¸…æ´—
                        day_df = day_df[day_df['å§“å'].notna() & day_df['é¢„è®¢äºº'].notna()]
                        
                        # é¢„è®¢äººå§“åæ ‡å‡†åŒ–å¤„ç†
                        def standardize_name(name):
                            if pd.isna(name):
                                return name
                            name_str = str(name).strip()
                            # å¤„ç†åŒä¹‰è¯
                            if name_str in ['å¹³å’Œ', 'å¹³å“¥']:
                                return 'å¹³å’Œ'
                            # å¤„ç†åˆ˜éœå’Œåˆ˜çš„æ˜ å°„
                            if name_str in ['åˆ˜éœ', 'åˆ˜']:
                                return 'åˆ˜éœ'
                            # å¤„ç†å‘¨å’Œå‘¨æ€ç—çš„æ˜ å°„
                            if name_str in ['å‘¨', 'å‘¨æ€ç—']:
                                return 'å‘¨æ€ç—'
                            # å¤„ç†å¤§å°å†™ç»Ÿä¸€ï¼ˆsk -> SKï¼‰
                            if name_str.lower() == 'sk':
                                return 'SK'
                            return name_str
                        
                        day_df['é¢„è®¢äºº'] = day_df['é¢„è®¢äºº'].apply(standardize_name)
                        
                        # é€‰æ‹©å’Œé‡å‘½ååˆ—
                        available_cols = ['æ—¥æœŸ', 'å¸‚åˆ«', 'åŒ…å¢', 'å§“å', 'é¢„è®¢äºº', 'ç»æ‰‹äºº']
                        existing_cols = [col for col in available_cols if col in day_df.columns]
                        day_df = day_df[existing_cols].copy()
                        
                        # æ ‡å‡†åŒ–åˆ—å
                        col_mapping = {'åŒ…å¢': 'æ¡Œç‰Œå·', 'å§“å': 'å®¢æˆ·å§“å'}
                        day_df.rename(columns=col_mapping, inplace=True)
                        
                        # å¤„ç†æ—¥æœŸ
                        if 'æ—¥æœŸ' in day_df.columns:
                            day_df['æ—¥æœŸ'] = pd.to_datetime(
                                day_df['æ—¥æœŸ'].astype(str).str.split().str[0], 
                                errors='coerce'
                            )
                        
                        # åˆå¹¶æ•°æ® - æ”¹è¿›çš„åŒ¹é…é€»è¾‘
                        if 'æ—¥æœŸ' in day_df.columns and 'æ¡Œç‰Œå·' in day_df.columns and 'å¸‚åˆ«' in day_df.columns:
                            # ä¸ºæ¯ä¸ªé¢„è®¢è®°å½•æ‰¾åˆ°æœ€ä½³åŒ¹é…çš„ç¾å›¢è®¢å•
                            merged_records = []
                            
                            for _, reservation in day_df.iterrows():
                                # æ‰¾åˆ°åŒä¸€æ—¥æœŸã€å¸‚åˆ«çš„æ‰€æœ‰ç¾å›¢è®¢å•ï¼Œç„¶åä½¿ç”¨æ™ºèƒ½æ¡Œç‰Œå·åŒ¹é…
                                reservation_date = reservation['æ—¥æœŸ'].date() if hasattr(reservation['æ—¥æœŸ'], 'date') else reservation['æ—¥æœŸ']
                                
                                # å…ˆæŒ‰æ—¥æœŸå’Œå¸‚åˆ«ç­›é€‰
                                candidate_orders = mt_df[
                                    (mt_df['ä¸‹å•æ—¥æœŸ'] == reservation_date) &
                                    (mt_df['å¸‚åˆ«'] == reservation['å¸‚åˆ«'])
                                ].copy()
                                
                                # ä½¿ç”¨æ™ºèƒ½åŒ¹é…æ‰¾åˆ°æ¡Œç‰Œå·åŒ¹é…çš„è®¢å•
                                matching_orders = []
                                match_info = []
                                
                                for _, order in candidate_orders.iterrows():
                                    is_match, match_type = self.smart_table_match(
                                        reservation['æ¡Œç‰Œå·'], 
                                        order['æ¡Œç‰Œå·']
                                    )
                                    if is_match:
                                        matching_orders.append(order)
                                        match_info.append(match_type)
                                
                                matching_orders = pd.DataFrame(matching_orders) if matching_orders else pd.DataFrame()
                                
                                if not matching_orders.empty:
                                    # ä¸ºæ¯ä¸ªåŒ¹é…çš„è®¢å•åˆ›å»ºç‹¬ç«‹è®°å½•
                                    for idx, (_, order) in enumerate(matching_orders.iterrows()):
                                        merged_record = reservation.copy()
                                        merged_record['æ”¯ä»˜åˆè®¡'] = order['æ”¯ä»˜åˆè®¡']
                                        merged_record['ä¸‹å•æ—¶é—´'] = order['ä¸‹å•æ—¶é—´']
                                        merged_record['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–'] = order['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–']
                                        merged_record['ç»“è´¦æ–¹å¼'] = order['ç»“è´¦æ–¹å¼']
                                        merged_record['åŒ¹é…ç±»å‹'] = match_info[idx] if idx < len(match_info) else 'æœªçŸ¥'
                                        merged_records.append(merged_record)
                                else:
                                    # æ²¡æœ‰åŒ¹é…çš„è®¢å•
                                    merged_record = reservation.copy()
                                    merged_record['æ”¯ä»˜åˆè®¡'] = None
                                    merged_record['ä¸‹å•æ—¶é—´'] = None
                                    merged_record['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–'] = None
                                    merged_record['ç»“è´¦æ–¹å¼'] = None
                                    merged_record['åŒ¹é…ç±»å‹'] = 'æœªåŒ¹é…'
                                    merged_records.append(merged_record)
                            
                            if merged_records:
                                merged = pd.DataFrame(merged_records)
                                merged_all = pd.concat([merged_all, merged], ignore_index=True)
                                
                    except Exception as e:
                        continue
            else:
                # å¦‚æœæ˜¯å•ä¸ªDataFrameï¼Œç›´æ¥å¤„ç†
                day_df = self.reservation_file.copy()
                
                # æ£€æŸ¥å¿…è¦çš„åˆ—æ˜¯å¦å­˜åœ¨
                required_cols = ['å§“å', 'é¢„è®¢äºº']
                missing_cols = [col for col in required_cols if col not in day_df.columns]
                if missing_cols:
                    return False, f"é¢„è®¢æ–‡ä»¶ç¼ºå°‘å¿…è¦åˆ—: {missing_cols}"
                    
                # æ•°æ®æ¸…æ´—
                day_df = day_df[day_df['å§“å'].notna() & day_df['é¢„è®¢äºº'].notna()]
                
                # é€‰æ‹©å’Œé‡å‘½ååˆ—
                available_cols = ['æ—¥æœŸ', 'å¸‚åˆ«', 'åŒ…å¢', 'å§“å', 'é¢„è®¢äºº', 'ç»æ‰‹äºº']
                existing_cols = [col for col in available_cols if col in day_df.columns]
                day_df = day_df[existing_cols].copy()
                
                # æ ‡å‡†åŒ–åˆ—å
                col_mapping = {'åŒ…å¢': 'æ¡Œç‰Œå·', 'å§“å': 'å®¢æˆ·å§“å'}
                day_df.rename(columns=col_mapping, inplace=True)
                
                # å¤„ç†æ—¥æœŸ
                if 'æ—¥æœŸ' in day_df.columns:
                    day_df['æ—¥æœŸ'] = pd.to_datetime(
                        day_df['æ—¥æœŸ'].astype(str).str.split().str[0], 
                        errors='coerce'
                    )
                
                # åˆå¹¶æ•°æ® - æ”¹è¿›çš„åŒ¹é…é€»è¾‘
                if 'æ—¥æœŸ' in day_df.columns and 'æ¡Œç‰Œå·' in day_df.columns and 'å¸‚åˆ«' in day_df.columns:
                    # ä¸ºæ¯ä¸ªé¢„è®¢è®°å½•æ‰¾åˆ°æœ€ä½³åŒ¹é…çš„ç¾å›¢è®¢å•
                    merged_records = []
                    
                    for _, reservation in day_df.iterrows():
                        # æ‰¾åˆ°åŒä¸€æ—¥æœŸã€æ¡Œç‰Œå·ã€å¸‚åˆ«çš„æ‰€æœ‰ç¾å›¢è®¢å•ï¼ˆä½¿ç”¨ä¸‹å•æ—¶é—´çš„æ—¥æœŸè¿›è¡ŒåŒ¹é…ï¼‰
                        reservation_date = reservation['æ—¥æœŸ'].date() if hasattr(reservation['æ—¥æœŸ'], 'date') else reservation['æ—¥æœŸ']
                        matching_orders = mt_df[
                            (mt_df['ä¸‹å•æ—¥æœŸ'] == reservation_date) &
                            (mt_df['æ¡Œç‰Œå·'] == reservation['æ¡Œç‰Œå·']) &
                            (mt_df['å¸‚åˆ«'] == reservation['å¸‚åˆ«'])
                        ].copy()
                        
                        if not matching_orders.empty:
                            # ä¸ºæ¯ä¸ªåŒ¹é…çš„è®¢å•åˆ›å»ºç‹¬ç«‹è®°å½•
                            for _, order in matching_orders.iterrows():
                                merged_record = reservation.copy()
                                merged_record['æ”¯ä»˜åˆè®¡'] = order['æ”¯ä»˜åˆè®¡']
                                merged_record['ä¸‹å•æ—¶é—´'] = order['ä¸‹å•æ—¶é—´']
                                merged_record['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–'] = order['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–']
                                merged_record['ç»“è´¦æ–¹å¼'] = order['ç»“è´¦æ–¹å¼']
                                merged_records.append(merged_record)
                        else:
                            # æ²¡æœ‰åŒ¹é…çš„è®¢å•
                            merged_record = reservation.copy()
                            merged_record['æ”¯ä»˜åˆè®¡'] = None
                            merged_record['ä¸‹å•æ—¶é—´'] = None
                            merged_record['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–'] = None
                            merged_record['ç»“è´¦æ–¹å¼'] = None
                            merged_records.append(merged_record)
                    
                    if merged_records:
                        merged_all = pd.DataFrame(merged_records)
            
            # æ•°æ®åå¤„ç†
            if not merged_all.empty:
                # æ·»åŠ åŒ¹é…çŠ¶æ€åˆ—
                merged_all['åŒ¹é…çŠ¶æ€'] = merged_all['æ”¯ä»˜åˆè®¡'].apply(
                    lambda x: 'å·²åŒ¹é…' if pd.notna(x) else 'æœªåŒ¹é…'
                )
                
                # æ ¼å¼åŒ–æ•°æ®
                if 'æ”¯ä»˜åˆè®¡' in merged_all.columns:
                    merged_all['æ”¯ä»˜åˆè®¡'] = merged_all['æ”¯ä»˜åˆè®¡'].apply(
                        lambda x: f"{x:.2f}" if pd.notna(x) else ""
                    )
                    
                # æ’åº
                sort_cols = []
                if 'æ—¥æœŸ' in merged_all.columns:
                    sort_cols.append('æ—¥æœŸ')
                if 'æ¡Œç‰Œå·' in merged_all.columns:
                    sort_cols.append('æ¡Œç‰Œå·')
                if sort_cols:
                    merged_all.sort_values(sort_cols, inplace=True, ignore_index=True)
            
            self.merged_df = merged_all
            self.original_df = merged_all.copy()  # ä¿å­˜åŸå§‹æ•°æ®
            
            # æ˜¾ç¤ºç»Ÿè®¡ä¿¡æ¯
            total_records = len(self.merged_df)
            matched_records = len(self.merged_df[self.merged_df['åŒ¹é…çŠ¶æ€'] == 'å·²åŒ¹é…']) if 'åŒ¹é…çŠ¶æ€' in self.merged_df.columns else 0
            
            return True, f"åŒ¹é…å®Œæˆï¼æ€»è®°å½•: {total_records}, å·²åŒ¹é…: {matched_records}, æœªåŒ¹é…: {total_records - matched_records}"
            
        except Exception as e:
            return False, f"åŒ¹é…å¤±è´¥: {str(e)}"
    
    def display_results(self):
        """æ˜¾ç¤ºåŒ¹é…ç»“æœ"""
        if self.merged_df.empty:
            st.warning("æš‚æ— æ•°æ®")
            return
        
        # æ˜¾ç¤ºåŒ¹é…ç»Ÿè®¡ä¿¡æ¯
        if 'åŒ¹é…ç±»å‹' in self.merged_df.columns:
            st.subheader("ğŸ“Š åŒ¹é…ç»Ÿè®¡")
            match_stats = self.merged_df['åŒ¹é…ç±»å‹'].value_counts()
            
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                complete_match = match_stats.get('å®Œå…¨åŒ¹é…', 0)
                st.metric("å®Œå…¨åŒ¹é…", complete_match, help="æ¡Œç‰Œå·å®Œå…¨ç›¸åŒçš„åŒ¹é…")
            
            with col2:
                number_match = match_stats.get('æ•°å­—åŒ¹é…', 0)
                st.metric("æ•°å­—åŒ¹é…", number_match, help="æ¡Œç‰Œå·æ•°å­—éƒ¨åˆ†ç›¸åŒçš„å ‚é£ŸåŒ¹é…")
            
            with col3:
                takeout_match = match_stats.get('å¤–å–åŒ¹é…', 0)
                st.metric("å¤–å–åŒ¹é…", takeout_match, help="é¢„è®¢æ”¹ä¸ºå¤–å–é…é€çš„åŒ¹é…")
            
            with col4:
                no_match = match_stats.get('æœªåŒ¹é…', 0)
                st.metric("æœªåŒ¹é…", no_match, help="æœªæ‰¾åˆ°å¯¹åº”ç¾å›¢è®¢å•")
            
            with col5:
                total_records = len(self.merged_df)
                match_rate = round((total_records - no_match) / total_records * 100, 1) if total_records > 0 else 0
                st.metric("åŒ¹é…ç‡", f"{match_rate}%", help="æˆåŠŸåŒ¹é…çš„è®°å½•æ¯”ä¾‹")
            
            st.divider()
        
        col1, col2 = st.columns(2)
        with col1:
            filter_option = st.selectbox(
                "æ˜¾ç¤ºå†…å®¹",
                ["å…¨éƒ¨è®°å½•", "å·²åŒ¹é…è®°å½•", "æœªåŒ¹é…è®°å½•"]
            )
            # ä¿å­˜ç­›é€‰æ¡ä»¶åˆ°session_state
            st.session_state.filter_option = filter_option
        
        with col2:
            search_keyword = st.text_input("æœç´¢é¢„è®¢äºº", placeholder="è¾“å…¥é¢„è®¢äººå§“åè¿›è¡Œæœç´¢")
            # ä¿å­˜æœç´¢å…³é”®è¯åˆ°session_state
            st.session_state.search_keyword = search_keyword
        
        # åº”ç”¨ç­›é€‰
        display_df = self.merged_df.copy()
        
        if filter_option == "å·²åŒ¹é…è®°å½•":
            display_df = display_df[display_df['åŒ¹é…çŠ¶æ€'] == 'å·²åŒ¹é…']
        elif filter_option == "æœªåŒ¹é…è®°å½•":
            display_df = display_df[display_df['åŒ¹é…çŠ¶æ€'] == 'æœªåŒ¹é…']
        
        if search_keyword:
            if 'é¢„è®¢äºº' in display_df.columns:
                # æ ‡å‡†åŒ–æœç´¢å…³é”®è¯
                def standardize_search_keyword(keyword):
                    keyword = keyword.strip()
                    if keyword in ['å¹³å’Œ', 'å¹³å“¥']:
                        return ['å¹³å’Œ', 'å¹³å“¥']  # è¿”å›æ‰€æœ‰åŒä¹‰è¯
                    elif keyword in ['åˆ˜éœ', 'åˆ˜']:
                        return ['åˆ˜éœ', 'åˆ˜']  # è¿”å›åˆ˜éœå’Œåˆ˜çš„æ‰€æœ‰å˜ä½“
                    elif keyword in ['å‘¨', 'å‘¨æ€ç—']:
                        return ['å‘¨', 'å‘¨æ€ç—']  # è¿”å›å‘¨å’Œå‘¨æ€ç—çš„æ‰€æœ‰å˜ä½“
                    elif keyword.lower() == 'sk':
                        return ['SK', 'sk', 'Sk', 'sK']  # è¿”å›æ‰€æœ‰å¤§å°å†™å˜ä½“
                    else:
                        return [keyword]
                
                search_terms = standardize_search_keyword(search_keyword)
                # åˆ›å»ºæœç´¢æ¡ä»¶ï¼ŒåŒ¹é…ä»»ä½•ä¸€ä¸ªåŒä¹‰è¯
                search_condition = False
                for term in search_terms:
                    search_condition |= display_df['é¢„è®¢äºº'].astype(str).str.contains(term, case=False, na=False)
                display_df = display_df[search_condition]
        
        # æ˜¾ç¤ºæ•°æ®è¡¨æ ¼ï¼ˆç®€åŒ–ç‰ˆï¼‰
        st.subheader(f"ğŸ“‹ æ•°æ®è¡¨æ ¼ ({len(display_df)} æ¡è®°å½•)")
        
        if not display_df.empty:
            # é…ç½®æ ¸å¿ƒåˆ—æ˜¾ç¤ºï¼ˆç®€åŒ–ä¿¡æ¯ï¼‰
            columns_to_show = ['æ—¥æœŸ', 'æ¡Œç‰Œå·', 'é¢„è®¢äºº', 'å¸‚åˆ«', 'åŒ¹é…çŠ¶æ€', 'åŒ¹é…ç±»å‹']
            available_columns = [col for col in columns_to_show if col in display_df.columns]
            
            # åˆ›å»ºæ˜¾ç¤ºç”¨çš„DataFrameå‰¯æœ¬å¹¶å¤„ç†æ•°æ®ç±»å‹
            table_df = display_df[available_columns].copy()
            
            # æ ¼å¼åŒ–æ˜¾ç¤º
            for col in table_df.columns:
                if col == 'åŒ¹é…çŠ¶æ€':
                    table_df[col] = table_df[col].apply(lambda x: 'âœ…å·²åŒ¹é…' if str(x) == 'å·²åŒ¹é…' else 'âŒæœªåŒ¹é…')
                else:
                    table_df[col] = table_df[col].astype(str).replace('nan', '')
            
            # é‡å‘½ååˆ—æ ‡é¢˜ä½¿å…¶æ›´ç®€æ´
            column_rename = {
                'æ—¥æœŸ': 'ğŸ“…æ—¥æœŸ',
                'æ¡Œç‰Œå·': 'ğŸª‘æ¡Œå·', 
                'é¢„è®¢äºº': 'ğŸ‘¤é¢„è®¢äºº',
                'å¸‚åˆ«': 'ğŸªå¸‚åˆ«',
                'åŒ¹é…çŠ¶æ€': 'ğŸ“ŠçŠ¶æ€'
            }
            table_df = table_df.rename(columns=column_rename)
            
            # æ·»åŠ æ°´å¹³æ»šåŠ¨çš„è¡¨æ ¼æ˜¾ç¤º
            st.markdown("""
            <style>
            .stDataFrame {
                overflow-x: auto;
            }
            .stDataFrame > div {
                overflow-x: auto;
            }
            </style>
            """, unsafe_allow_html=True)
            
            # ä½¿ç”¨å¯é€‰æ‹©çš„æ•°æ®è¡¨æ ¼
            selected_rows = st.dataframe(
                table_df,
                use_container_width=True,
                height=None,  # ç§»é™¤é«˜åº¦é™åˆ¶ï¼Œæ˜¾ç¤ºæ‰€æœ‰å†…å®¹
                on_select="rerun",
                selection_mode="single-row"
            )
            
            # å¤„ç†è¡Œé€‰æ‹©å’Œè¯¦æƒ…æ˜¾ç¤º
            if selected_rows.selection.rows:
                selected_idx = selected_rows.selection.rows[0]
                if selected_idx < len(display_df):
                    selected_record = display_df.iloc[selected_idx]
                    self.show_record_details(selected_record, display_df, selected_idx)
            
            # æ‰‹åŠ¨åŒ¹é…åŠŸèƒ½
            if filter_option == "æœªåŒ¹é…è®°å½•" and not display_df.empty:
                self.manual_match_interface(display_df)
        else:
            st.info("ğŸ“ æ²¡æœ‰ç¬¦åˆæ¡ä»¶çš„è®°å½•")
    
    def manual_match_interface(self, unmatched_df):
        """æ‰‹åŠ¨åŒ¹é…ç•Œé¢"""
        st.write("**ğŸ”§ æ‰‹åŠ¨åŒ¹é…**")
        
        if unmatched_df.empty:
            return
        
        # é€‰æ‹©è¦åŒ¹é…çš„é¢„è®¢è®°å½•ï¼ˆç®€åŒ–æ˜¾ç¤ºï¼‰
        reservation_options = []
        for idx, row in unmatched_df.iterrows():
            # æ ¼å¼åŒ–æ—¥æœŸï¼Œåªæ˜¾ç¤ºå¹´æœˆæ—¥
            date_str = 'N/A'
            if pd.notna(row.get('æ—¥æœŸ')):
                try:
                    if hasattr(row.get('æ—¥æœŸ'), 'strftime'):
                        date_str = row.get('æ—¥æœŸ').strftime('%Y-%m-%d')
                    else:
                        date_str = str(row.get('æ—¥æœŸ')).split(' ')[0]  # å–ç©ºæ ¼å‰çš„æ—¥æœŸéƒ¨åˆ†
                except:
                    date_str = str(row.get('æ—¥æœŸ', 'N/A'))
            
            option_text = f"ğŸ“…{date_str} | ğŸª‘{row.get('æ¡Œç‰Œå·', 'N/A')}æ¡Œ | ğŸª{row.get('å¸‚åˆ«', 'N/A')} | ğŸ‘¤{row.get('é¢„è®¢äºº', 'N/A')}"
            reservation_options.append((option_text, idx))
        
        selected_reservation = st.selectbox(
            "é€‰æ‹©è¦åŒ¹é…çš„é¢„è®¢è®°å½•",
            options=reservation_options,
            format_func=lambda x: x[0]
        )
        
        if selected_reservation and self.meituan_file is not None:
            reservation_idx = selected_reservation[1]
            reservation_record = unmatched_df.loc[reservation_idx]
            
            # è·å–ç›¸å…³çš„ç¾å›¢è®¢å•
            reservation_date = reservation_record['æ—¥æœŸ']
            if hasattr(reservation_date, 'date'):
                reservation_date = reservation_date.date()
            
            # å¤„ç†ç¾å›¢æ•°æ®ï¼Œç¡®ä¿æ”¯ä»˜åˆè®¡å­—æ®µæ­£ç¡®æå–
            def extract_payment(payment_str):
                if pd.isna(payment_str):
                    return None
                try:
                    # æŸ¥æ‰¾æ‰€æœ‰æ•°å­—ï¼ˆåŒ…æ‹¬è´Ÿæ•°å’Œå°æ•°ï¼‰
                    numbers = re.findall(r'-?\d+\.?\d*', str(payment_str))
                    if numbers:
                        return float(numbers[0])
                except (ValueError, IndexError):
                    pass
                return None
            
            # å¤åˆ¶ç¾å›¢æ–‡ä»¶å¹¶å¤„ç†æ”¯ä»˜åˆè®¡
            meituan_processed = self.meituan_file.copy()
            if 'æ”¯ä»˜åˆè®¡' not in meituan_processed.columns and 'ç»“è´¦æ–¹å¼' in meituan_processed.columns:
                meituan_processed['æ”¯ä»˜åˆè®¡'] = meituan_processed['ç»“è´¦æ–¹å¼'].apply(extract_payment)
            
            # å®‰å…¨åœ°æ¯”è¾ƒæ—¥æœŸï¼ˆä½¿ç”¨ä¸‹å•æ—¶é—´çš„æ—¥æœŸè¿›è¡ŒåŒ¹é…ï¼‰
            try:
                if 'ä¸‹å•æ—¶é—´' in meituan_processed.columns:
                    # ç¡®ä¿ä¸‹å•æ—¶é—´æ˜¯datetimeç±»å‹
                    meituan_dates = pd.to_datetime(meituan_processed['ä¸‹å•æ—¶é—´'], errors='coerce')
                    related_meituan = meituan_processed[
                        meituan_dates.dt.date == reservation_date
                    ]
                else:
                    related_meituan = meituan_processed
            except (AttributeError, TypeError):
                related_meituan = meituan_processed
            
            if related_meituan.empty:
                related_meituan = meituan_processed
            
            st.write("**ğŸ“‹ å¯é€‰æ‹©çš„ç¾å›¢è®¢å•:**")
            st.write("ğŸ’¡ *ç‚¹å‡»è¡¨æ ¼ä¸­çš„è¡Œæ¥é€‰æ‹©ç¾å›¢è®¢å•ï¼ˆæ”¯æŒå¤šé€‰ï¼ŒæŒ‰ä½Ctrlé”®å¯é€‰æ‹©å¤šä¸ªï¼‰*")
            
            # æ˜¾ç¤ºç¾å›¢è®¢å•è¯¦ç»†ä¿¡æ¯è¡¨æ ¼ï¼ˆå¯é€‰æ‹©ï¼‰
            if not related_meituan.empty:
                # é€‰æ‹©è¦æ˜¾ç¤ºçš„æ ¸å¿ƒåˆ—ï¼ˆåŒ…å«æ—¥æœŸã€æ—¶é—´ã€é‡‘é¢ï¼‰
                display_columns = ['è¥ä¸šæ—¥æœŸ', 'æ¡Œç‰Œå·', 'ä¸‹å•æ—¶é—´', 'æ”¯ä»˜åˆè®¡', 'ç»“è´¦æ–¹å¼']
                available_columns = [col for col in display_columns if col in related_meituan.columns]
                
                if available_columns:
                    meituan_display = related_meituan[available_columns].copy()
                    # æ ¼å¼åŒ–æ˜¾ç¤º
                    for col in meituan_display.columns:
                        if col == 'æ”¯ä»˜åˆè®¡':
                            meituan_display[col] = meituan_display[col].apply(lambda x: f"Â¥{x}" if pd.notna(x) and str(x) != 'nan' else '')
                        else:
                            meituan_display[col] = meituan_display[col].astype(str).replace('nan', '')
                    
                    # é‡å‘½ååˆ—æ ‡é¢˜ä½¿å…¶æ›´ç®€æ´
                    column_rename = {
                        'è¥ä¸šæ—¥æœŸ': 'ğŸ“…è¥ä¸šæ—¥æœŸ',
                        'æ¡Œç‰Œå·': 'ğŸª‘æ¡Œå·',
                        'ä¸‹å•æ—¶é—´': 'â°ä¸‹å•æ—¶é—´', 
                        'æ”¯ä»˜åˆè®¡': 'ğŸ’°æ”¯ä»˜åˆè®¡',
                        'ç»“è´¦æ–¹å¼': 'ğŸ’³ç»“è´¦æ–¹å¼'
                    }
                    meituan_display = meituan_display.rename(columns=column_rename)
                    
                    # ä½¿ç”¨å¯é€‰æ‹©çš„æ•°æ®æ¡†ï¼ˆæ”¯æŒå¤šé€‰ï¼‰
                    selected_rows = st.dataframe(
                        meituan_display,
                        use_container_width=True,
                        height=200,
                        hide_index=False,
                        selection_mode="multi-row",
                        key="meituan_selector",
                        on_select="rerun"
                    )
                    
                    # è·å–é€‰ä¸­çš„è¡Œï¼ˆæ”¯æŒå¤šé€‰ï¼‰
                    selected_meituan_indices = []
                    if selected_rows and 'selection' in selected_rows and 'rows' in selected_rows['selection']:
                        if selected_rows['selection']['rows']:
                            for selected_row_idx in selected_rows['selection']['rows']:
                                actual_idx = related_meituan.index[selected_row_idx]
                                selected_meituan_indices.append(actual_idx)
                    
                    # æ˜¾ç¤ºé€‰ä¸­çš„è®¢å•æ•°é‡
                    if selected_meituan_indices:
                        st.info(f"å·²é€‰æ‹© {len(selected_meituan_indices)} ä¸ªç¾å›¢è®¢å•")
            
            # ç¡®è®¤åŒ¹é…æŒ‰é’®
            if st.button("ç¡®è®¤åŒ¹é…", type="primary"):
                if selected_meituan_indices:
                    # è·å–åŸå§‹é¢„è®¢è®°å½•
                    original_reservation = self.merged_df.loc[reservation_idx].copy()
                    
                    # ä¸ºæ¯ä¸ªé€‰ä¸­çš„ç¾å›¢è®¢å•åˆ›å»ºåŒ¹é…è®°å½•
                    new_records = []
                    for i, meituan_idx in enumerate(selected_meituan_indices):
                        meituan_record = related_meituan.loc[meituan_idx]
                        
                        # åˆ›å»ºæ–°çš„åŒ¹é…è®°å½•
                        new_record = original_reservation.copy()
                        new_record['åŒ¹é…çŠ¶æ€'] = 'å·²åŒ¹é…'
                        new_record['ä¸‹å•æ—¶é—´'] = str(meituan_record.get('ä¸‹å•æ—¶é—´', ''))
                        new_record['ä¸‹å•æ—¶é—´_æ ¼å¼åŒ–'] = str(meituan_record.get('ä¸‹å•æ—¶é—´', ''))
                        new_record['ç»“è´¦æ–¹å¼'] = str(meituan_record.get('ç»“è´¦æ–¹å¼', ''))
                        
                        # å¦‚æœæ˜¯ç¬¬ä¸€ä¸ªè®°å½•ï¼Œæ›´æ–°åŸè®°å½•ï¼›å¦åˆ™æ·»åŠ æ–°è®°å½•
                        if i == 0:
                            # æ›´æ–°åŸè®°å½•
                            for col in new_record.index:
                                self.merged_df.at[reservation_idx, col] = new_record[col]
                        else:
                            # æ·»åŠ æ–°è®°å½•åˆ°åˆ—è¡¨
                            new_records.append(new_record)
                    
                    # å°†æ–°è®°å½•æ·»åŠ åˆ°DataFrame
                    if new_records:
                        new_df = pd.DataFrame(new_records)
                        self.merged_df = pd.concat([self.merged_df, new_df], ignore_index=True)
                    
                    st.success(f"åŒ¹é…æˆåŠŸï¼å·²ä¸º {len(selected_meituan_indices)} ä¸ªç¾å›¢è®¢å•åˆ›å»ºåŒ¹é…è®°å½•ã€‚é¡µé¢å°†è‡ªåŠ¨åˆ·æ–°")
                    st.rerun()
                else:
                    st.warning("è¯·å…ˆé€‰æ‹©è¦åŒ¹é…çš„ç¾å›¢è®¢å•")
    
    def export_results(self):
        """å¯¼å‡ºç»“æœ"""
        if self.merged_df.empty:
            st.warning("æš‚æ— æ•°æ®")
            return
        
        # å¯¼å‡ºé€‰é¡¹
        export_option = st.selectbox(
            "å¯¼å‡ºé€‰é¡¹",
            ["ä»…æœç´¢", "å…¨éƒ¨ï¼ˆæŒ‰æ—¶é—´æ’åˆ—ï¼‰"]
        )
        
        # è·å–å½“å‰æœç´¢å’Œç­›é€‰æ¡ä»¶
        if 'filter_option' not in st.session_state:
            st.session_state.filter_option = "å…¨éƒ¨è®°å½•"
        if 'search_keyword' not in st.session_state:
            st.session_state.search_keyword = ""
        
        # å‡†å¤‡å¯¼å‡ºæ•°æ®
        if export_option == "ä»…æœç´¢":
            # è·å–å½“å‰æ˜¾ç¤ºçš„æœç´¢ç»“æœï¼ˆåªåŒ…å«åŒ¹é…æˆåŠŸçš„ï¼‰
            export_df = self.get_filtered_data()
            export_df = export_df[export_df['åŒ¹é…çŠ¶æ€'] == 'å·²åŒ¹é…']  # åªå¯¼å‡ºåŒ¹é…æˆåŠŸçš„
            
            # æ ¹æ®æœç´¢å…³é”®è¯ç”Ÿæˆæ–‡ä»¶å
            search_keyword = getattr(st.session_state, 'search_keyword', "")
            if search_keyword.strip():
                filename_suffix = f"{search_keyword.strip()}ï¼ˆç¾å›¢åŒ¹é…æ¸…å•ï¼‰"
            else:
                filename_suffix = "æœç´¢ç»“æœ"
        else:
            # å…¨éƒ¨åŒ¹é…æˆåŠŸçš„æ•°æ®ï¼ŒæŒ‰æ—¶é—´æ’åˆ—
            export_df = self.merged_df[self.merged_df['åŒ¹é…çŠ¶æ€'] == 'å·²åŒ¹é…'].copy()
            if 'æ—¥æœŸ' in export_df.columns:
                export_df = export_df.sort_values('æ—¥æœŸ')
            filename_suffix = "å…¨éƒ¨åŒ¹é…"
        
        if export_df.empty:
            st.warning("æ²¡æœ‰åŒ¹é…æˆåŠŸçš„æ•°æ®å¯å¯¼å‡º")
            return
        
        # å‡†å¤‡å¯¼å‡ºçš„åˆ—
        export_columns = ['ä¸‹å•æ—¶é—´', 'é¢„è®¢äºº', 'æ¡Œç‰Œå·', 'æ”¯ä»˜åˆè®¡', 'ç»“è´¦æ–¹å¼', 'åŒ¹é…ç±»å‹']
        
        # æ£€æŸ¥å¹¶é€‰æ‹©å¯ç”¨çš„åˆ—
        available_columns = []
        for col in export_columns:
            if col in export_df.columns:
                available_columns.append(col)
            elif col == 'é¢„è®¢äºº' and 'å®¢æˆ·å§“å' in export_df.columns:
                available_columns.append('å®¢æˆ·å§“å')
                export_df = export_df.rename(columns={'å®¢æˆ·å§“å': 'é¢„è®¢äºº'})
        
        # åˆ›å»ºå¯¼å‡ºç”¨çš„DataFrame
        final_export_df = export_df[available_columns].copy()
        
        # æŒ‰æ—¥æœŸæ’åºï¼ˆå¦‚æœæœ‰ä¸‹å•æ—¶é—´åˆ—ï¼‰
        if 'ä¸‹å•æ—¶é—´' in final_export_df.columns:
            final_export_df = final_export_df.sort_values('ä¸‹å•æ—¶é—´')
        
        # åˆ›å»ºExcelæ–‡ä»¶
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            final_export_df.to_excel(writer, sheet_name='åŒ¹é…ç»“æœ', index=False)
            
            # è·å–å·¥ä½œè¡¨å¹¶è®¾ç½®æ ¼å¼
            worksheet = writer.sheets['åŒ¹é…ç»“æœ']
            
            # è®¾ç½®Excelæ ¼å¼
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # å®šä¹‰æ ·å¼
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # è®¾ç½®è¡¨å¤´æ ·å¼
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # è®¾ç½®æ•°æ®è¡Œæ ·å¼
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = border
            
            # æ™ºèƒ½è°ƒæ•´åˆ—å®½
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # è®¡ç®—åˆ—çš„æœ€å¤§å†…å®¹é•¿åº¦
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        # ä¸­æ–‡å­—ç¬¦æŒ‰2ä¸ªå­—ç¬¦è®¡ç®—å®½åº¦
                        char_count = sum(2 if ord(char) > 127 else 1 for char in cell_value)
                        if char_count > max_length:
                            max_length = char_count
                    except:
                        pass
                
                # æ ¹æ®åˆ—å†…å®¹è®¾ç½®åˆé€‚çš„å®½åº¦
                if column_letter == 'A':  # ä¸‹å•æ—¶é—´åˆ—
                    adjusted_width = max(22, min(max_length + 4, 28))
                elif column_letter == 'B':  # é¢„è®¢äººåˆ—
                    adjusted_width = max(15, min(max_length + 3, 25))
                elif column_letter == 'C':  # æ¡Œç‰Œå·åˆ—
                    adjusted_width = max(12, min(max_length + 3, 18))
                elif column_letter == 'D':  # æ”¯ä»˜åˆè®¡åˆ—
                    adjusted_width = max(15, min(max_length + 3, 22))
                elif column_letter == 'E':  # ç»“è´¦æ–¹å¼åˆ—
                    adjusted_width = max(25, min(max_length + 5, 40))  # å¢åŠ ç»“è´¦æ–¹å¼åˆ—å®½åº¦
                elif column_letter == 'F':  # åŒ¹é…ç±»å‹åˆ—
                    adjusted_width = max(12, min(max_length + 3, 18))
                else:
                    adjusted_width = max(15, min(max_length + 3, 35))
                
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # è®¾ç½®è¡Œé«˜
            for row in range(1, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 35  # å¢åŠ è¡Œé«˜ä»¥é€‚åº”å¤šè¡Œå†…å®¹
            
            # ç‰¹åˆ«å¤„ç†è¡¨å¤´è¡Œé«˜
            worksheet.row_dimensions[1].height = 30
        
        excel_data = output.getvalue()
        
        # ç”Ÿæˆæ–‡ä»¶å
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"åŒ¹é…ç»“æœ_{filename_suffix}_{timestamp}.xlsx"
        
        st.download_button(
            label=f"ğŸ“¥ ä¸‹è½½Excel ({len(final_export_df)}æ¡è®°å½•)",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    def get_filtered_data(self):
        """è·å–å½“å‰ç­›é€‰å’Œæœç´¢åçš„æ•°æ®"""
        display_df = self.merged_df.copy()
        
        # åº”ç”¨ç­›é€‰ï¼ˆä»session_stateè·å–å½“å‰ç­›é€‰æ¡ä»¶ï¼‰
        filter_option = getattr(st.session_state, 'filter_option', "å…¨éƒ¨è®°å½•")
        search_keyword = getattr(st.session_state, 'search_keyword', "")
        
        if filter_option == "å·²åŒ¹é…è®°å½•":
            display_df = display_df[display_df['åŒ¹é…çŠ¶æ€'] == 'å·²åŒ¹é…']
        elif filter_option == "æœªåŒ¹é…è®°å½•":
            display_df = display_df[display_df['åŒ¹é…çŠ¶æ€'] == 'æœªåŒ¹é…']
        
        if search_keyword:
            if 'é¢„è®¢äºº' in display_df.columns:
                # æ ‡å‡†åŒ–æœç´¢å…³é”®è¯
                def standardize_search_keyword(keyword):
                    keyword = keyword.strip()
                    if keyword in ['å¹³å’Œ', 'å¹³å“¥']:
                        return ['å¹³å’Œ', 'å¹³å“¥']  # è¿”å›æ‰€æœ‰åŒä¹‰è¯
                    elif keyword in ['åˆ˜éœ', 'åˆ˜']:
                        return ['åˆ˜éœ', 'åˆ˜']  # è¿”å›åˆ˜éœå’Œåˆ˜çš„æ‰€æœ‰å˜ä½“
                    elif keyword in ['å‘¨', 'å‘¨æ€ç—']:
                        return ['å‘¨', 'å‘¨æ€ç—']  # è¿”å›å‘¨å’Œå‘¨æ€ç—çš„æ‰€æœ‰å˜ä½“
                    elif keyword.lower() == 'sk':
                        return ['SK', 'sk', 'Sk', 'sK']  # è¿”å›æ‰€æœ‰å¤§å°å†™å˜ä½“
                    else:
                        return [keyword]
                
                search_terms = standardize_search_keyword(search_keyword)
                # åˆ›å»ºæœç´¢æ¡ä»¶ï¼ŒåŒ¹é…ä»»ä½•ä¸€ä¸ªåŒä¹‰è¯
                search_condition = False
                for term in search_terms:
                    search_condition |= display_df['é¢„è®¢äºº'].astype(str).str.contains(term, case=False, na=False)
                display_df = display_df[search_condition]
        
        return display_df
    
    def normalize_customer_name(self, name):
        """æ ‡å‡†åŒ–é¢„è®¢äººå§“å"""
        if pd.isna(name) or str(name).strip() == '':
            return None
        
        name = str(name).strip()
        
        # è½¬æ¢ä¸ºå°å†™è¿›è¡Œæ¯”è¾ƒ
        name_lower = name.lower()
        
        # å®šä¹‰å§“åæ˜ å°„è§„åˆ™
        name_mappings = {
            'sk': 'SK',  # sk -> SK
            'å¹³': 'å¹³å“¥',  # å¹³ -> å¹³å“¥
            'å¹³å“¥': 'å¹³å“¥',  # å¹³å“¥ä¿æŒä¸å˜
            'å‘¨': 'å‘¨',  # å‘¨ä¿æŒä¸å˜
        }
        
        # æ£€æŸ¥æ˜¯å¦éœ€è¦æ˜ å°„
        for key, value in name_mappings.items():
            if name_lower == key.lower():
                return value
        
        # å¦‚æœæ²¡æœ‰ç‰¹æ®Šæ˜ å°„ï¼Œè¿”å›åŸå§‹åç§°ï¼ˆä¿æŒåŸæœ‰å¤§å°å†™ï¼‰
        return name
    
    def get_standardized_customers(self):
        """è·å–æ ‡å‡†åŒ–åçš„é¢„è®¢äººåˆ—è¡¨"""
        if 'é¢„è®¢äºº' not in self.merged_df.columns:
            return []
        
        # æ ‡å‡†åŒ–æ‰€æœ‰é¢„è®¢äººå§“å
        standardized_names = self.merged_df['é¢„è®¢äºº'].apply(self.normalize_customer_name)
        standardized_names = standardized_names.dropna().unique()
        
        return sorted([name for name in standardized_names if name])
    
    def show_data_analysis(self):
        """æ˜¾ç¤ºæ•°æ®åˆ†æé¡µé¢"""
        st.header("ğŸ“ˆ é¢„è®¢äººæ•°æ®åˆ†æ")
        
        if self.merged_df.empty:
            st.warning("æš‚æ— æ•°æ®ï¼Œè¯·å…ˆåœ¨'æ–‡ä»¶å¤„ç†'æ ‡ç­¾é¡µä¸­ä¸Šä¼ æ–‡ä»¶å¹¶è¿›è¡ŒåŒ¹é…")
            return
        
        # åˆ›å»ºä¸¤åˆ—å¸ƒå±€
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.subheader("ğŸ” é¢„è®¢äººæœç´¢")
            
            # è·å–æ ‡å‡†åŒ–åçš„é¢„è®¢äººåˆ—è¡¨
            if 'é¢„è®¢äºº' in self.merged_df.columns:
                all_customers = self.get_standardized_customers()
                
                # æœç´¢æ¡†
                search_customer = st.selectbox(
                    "é€‰æ‹©é¢„è®¢äºº",
                    options=["è¯·é€‰æ‹©..."] + all_customers,
                    key="customer_analysis_search"
                )
                
                # æˆ–è€…è¾“å…¥æœç´¢
                manual_search = st.text_input(
                    "æˆ–æ‰‹åŠ¨è¾“å…¥é¢„è®¢äººå§“å",
                    placeholder="è¾“å…¥é¢„è®¢äººå§“åè¿›è¡Œæœç´¢",
                    key="manual_customer_search"
                )
                
                # ç¡®å®šæœ€ç»ˆæœç´¢çš„å®¢æˆ·
                target_customer = None
                if manual_search.strip():
                    target_customer = manual_search.strip()
                elif search_customer != "è¯·é€‰æ‹©...":
                    target_customer = search_customer
                
                if target_customer:
                    st.success(f"å·²é€‰æ‹©ï¼š{target_customer}")
                    
                    # åˆ†ææŒ‰é’®
                    if st.button("ğŸ“Š å¼€å§‹åˆ†æ", type="primary", use_container_width=True):
                        st.session_state.analysis_customer = target_customer
                        st.rerun()
            else:
                st.error("æ•°æ®ä¸­æœªæ‰¾åˆ°'é¢„è®¢äºº'å­—æ®µ")
        
        with col2:
            st.subheader("ğŸ“Š åˆ†æç»“æœ")
            
            # æ£€æŸ¥æ˜¯å¦æœ‰é€‰æ‹©çš„å®¢æˆ·è¿›è¡Œåˆ†æ
            if hasattr(st.session_state, 'analysis_customer') and st.session_state.analysis_customer:
                customer_name = st.session_state.analysis_customer
                
                # ç­›é€‰è¯¥å®¢æˆ·çš„æ•°æ®ï¼ˆä½¿ç”¨æ ‡å‡†åŒ–å§“ååŒ¹é…ï¼‰
                standardized_customer_names = self.merged_df['é¢„è®¢äºº'].apply(self.normalize_customer_name)
                customer_data = self.merged_df[standardized_customer_names == customer_name]
                
                if customer_data.empty:
                    st.warning(f"æœªæ‰¾åˆ°é¢„è®¢äºº'{customer_name}'çš„ç›¸å…³æ•°æ®")
                else:
                    # æ˜¾ç¤ºåŸºæœ¬ç»Ÿè®¡ä¿¡æ¯
                    st.markdown(f"### ğŸ‘¤ {customer_name} çš„é¢„è®¢ç»Ÿè®¡")
                    
                    # åˆ›å»ºæŒ‡æ ‡å¡ç‰‡
                    metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
                    
                    with metric_col1:
                        total_orders = len(customer_data)
                        st.metric("æ€»é¢„è®¢æ¬¡æ•°", total_orders)
                    
                    with metric_col2:
                        matched_orders = len(customer_data[customer_data['åŒ¹é…çŠ¶æ€'] == 'å·²åŒ¹é…'])
                        st.metric("æˆåŠŸåŒ¹é…", matched_orders)
                    
                    with metric_col3:
                        if matched_orders > 0:
                            match_rate = round((matched_orders / total_orders) * 100, 1)
                            st.metric("åŒ¹é…ç‡", f"{match_rate}%")
                        else:
                            st.metric("åŒ¹é…ç‡", "0%")
                    
                    with metric_col4:
                        # è®¡ç®—æ€»æ¶ˆè´¹é‡‘é¢ï¼ˆä»…åŒ¹é…æˆåŠŸçš„è®¢å•ï¼‰
                        matched_data = customer_data[customer_data['åŒ¹é…çŠ¶æ€'] == 'å·²åŒ¹é…']
                        if not matched_data.empty and 'æ”¯ä»˜åˆè®¡' in matched_data.columns:
                            # æå–æ”¯ä»˜é‡‘é¢çš„æ•°å­—éƒ¨åˆ†
                            amounts = matched_data['æ”¯ä»˜åˆè®¡'].astype(str).str.extract(r'([0-9.]+)').astype(float)
                            total_amount = amounts.sum().iloc[0] if not amounts.empty else 0
                            st.metric("æ€»æ¶ˆè´¹é‡‘é¢", f"Â¥{total_amount:.2f}")
                        else:
                            st.metric("æ€»æ¶ˆè´¹é‡‘é¢", "Â¥0.00")
                    
                    st.divider()
                    
                    # å¯è§†åŒ–å›¾è¡¨
                    chart_col1, chart_col2 = st.columns(2)
                    
                    with chart_col1:
                        # å·¥ä½œæ—¥vså‘¨æœ«åˆ†æï¼ˆä»…åˆ†æå·²åŒ¹é…æ•°æ®ï¼‰
                        st.markdown("#### ğŸ“… å·¥ä½œæ—¥vså‘¨æœ«åˆ†æ")
                        if 'æ—¥æœŸ' in customer_data.columns:
                            # åªåˆ†æå·²åŒ¹é…çš„æ•°æ®
                            matched_data = customer_data[customer_data['åŒ¹é…çŠ¶æ€'] == 'å·²åŒ¹é…'].copy()
                            
                            if not matched_data.empty:
                                # è½¬æ¢æ—¥æœŸæ ¼å¼å¹¶åˆ†æå·¥ä½œæ—¥/å‘¨æœ«
                                matched_data['æ—¥æœŸ'] = pd.to_datetime(matched_data['æ—¥æœŸ'], errors='coerce')
                                customer_data_copy = matched_data.dropna(subset=['æ—¥æœŸ'])
                            else:
                                customer_data_copy = pd.DataFrame()  # ç©ºæ•°æ®æ¡†
                            
                            if not customer_data_copy.empty:
                                # æ·»åŠ æ˜ŸæœŸå‡ åˆ— (0=å‘¨ä¸€, 6=å‘¨æ—¥)
                                customer_data_copy['weekday'] = customer_data_copy['æ—¥æœŸ'].dt.dayofweek
                                customer_data_copy['day_type'] = customer_data_copy['weekday'].apply(
                                    lambda x: 'å‘¨æœ«' if x >= 5 else 'å·¥ä½œæ—¥'
                                )
                                
                                # ç»Ÿè®¡å·¥ä½œæ—¥vså‘¨æœ«çš„é¢„è®¢æ¬¡æ•°
                                day_type_counts = customer_data_copy['day_type'].value_counts()
                                
                                if not day_type_counts.empty:
                                    fig_daytype = px.bar(
                                        x=day_type_counts.index,
                                        y=day_type_counts.values,
                                        title=f"{customer_name} çš„å·¥ä½œæ—¥vså‘¨æœ«é¢„è®¢åˆ†å¸ƒï¼ˆå·²åŒ¹é…æ•°æ®ï¼‰",
                                        labels={'x': 'æ—¥æœŸç±»å‹', 'y': 'é¢„è®¢æ¬¡æ•°'},
                                        color=day_type_counts.index,
                                        color_discrete_map={
                                            'å·¥ä½œæ—¥': '#3b82f6',
                                            'å‘¨æœ«': '#f59e0b'
                                        }
                                    )
                                    fig_daytype.update_layout(
                                        height=300,
                                        font=dict(family="Microsoft YaHei, SimHei, sans-serif"),
                                        showlegend=False
                                    )
                                    st.plotly_chart(fig_daytype, use_container_width=True)
                                    
                                    # æ˜¾ç¤ºç»Ÿè®¡ä¿¡æ¯
                                    workday_count = day_type_counts.get('å·¥ä½œæ—¥', 0)
                                    weekend_count = day_type_counts.get('å‘¨æœ«', 0)
                                    total_count = workday_count + weekend_count
                                    
                                    if total_count > 0:
                                        workday_pct = round((workday_count / total_count) * 100, 1)
                                        weekend_pct = round((weekend_count / total_count) * 100, 1)
                                        
                                        st.markdown(f"""
                                        **ğŸ“Š ç»Ÿè®¡æ‘˜è¦ï¼ˆå·²åŒ¹é…æ•°æ®ï¼‰ï¼š**
                                        - å·¥ä½œæ—¥é¢„è®¢ï¼š{workday_count}æ¬¡ ({workday_pct}%)
                                        - å‘¨æœ«é¢„è®¢ï¼š{weekend_count}æ¬¡ ({weekend_pct}%)
                                        """)
                                else:
                                    st.info("æš‚æ— æœ‰æ•ˆçš„å·²åŒ¹é…æ—¥æœŸæ•°æ®è¿›è¡Œå·¥ä½œæ—¥/å‘¨æœ«åˆ†æ")
                            else:
                                st.info("è¯¥é¢„è®¢äººæš‚æ— å·²åŒ¹é…çš„æ•°æ®ï¼Œæ— æ³•è¿›è¡Œå·¥ä½œæ—¥/å‘¨æœ«åˆ†æ")
                        else:
                            st.info("æ•°æ®ä¸­ç¼ºå°‘æ—¥æœŸå­—æ®µï¼Œæ— æ³•è¿›è¡Œå·¥ä½œæ—¥/å‘¨æœ«åˆ†æ")
                    
                    with chart_col2:
                        # é¢„è®¢æ—¶é—´è¶‹åŠ¿å›¾
                        st.markdown("#### ğŸ“… é¢„è®¢æ—¶é—´è¶‹åŠ¿")
                        if 'æ—¥æœŸ' in customer_data.columns:
                            # æŒ‰æ—¥æœŸç»Ÿè®¡é¢„è®¢æ¬¡æ•°
                            date_counts = customer_data['æ—¥æœŸ'].value_counts().sort_index()
                            
                            if not date_counts.empty:
                                 fig_line = px.line(
                                     x=date_counts.index,
                                     y=date_counts.values,
                                     title=f"{customer_name} çš„é¢„è®¢æ—¶é—´è¶‹åŠ¿",
                                     labels={'x': 'æ—¥æœŸ', 'y': 'é¢„è®¢æ¬¡æ•°'}
                                 )
                                 fig_line.update_layout(
                                      height=300,
                                      font=dict(family="Microsoft YaHei, SimHei, sans-serif"),
                                      xaxis=dict(
                                          tickformat='%Yå¹´%mæœˆ%dæ—¥',
                                          tickangle=45
                                      )
                                  )
                                 st.plotly_chart(fig_line, use_container_width=True)
                            else:
                                st.info("æš‚æ— æ—¥æœŸæ•°æ®")
                        else:
                            st.info("æ•°æ®ä¸­æœªåŒ…å«æ—¥æœŸä¿¡æ¯")
                    
                    # æ¡Œç‰Œå·åå¥½åˆ†æ
                    if 'æ¡Œç‰Œå·' in customer_data.columns:
                        st.markdown("#### ğŸª‘ æ¡Œç‰Œå·åå¥½åˆ†æ")
                        table_counts = customer_data['æ¡Œç‰Œå·'].value_counts().head(10)
                        
                        if not table_counts.empty:
                             fig_bar = px.bar(
                                 x=table_counts.values,
                                 y=table_counts.index,
                                 orientation='h',
                                 title=f"{customer_name} çš„æ¡Œç‰Œå·åå¥½ (å‰10)",
                                 labels={'x': 'é¢„è®¢æ¬¡æ•°', 'y': 'æ¡Œç‰Œå·'}
                             )
                             fig_bar.update_layout(
                                 height=400,
                                 font=dict(family="Microsoft YaHei, SimHei, sans-serif")
                             )
                             st.plotly_chart(fig_bar, use_container_width=True)
                    
                    # è¯¦ç»†æ•°æ®è¡¨æ ¼
                    st.markdown("#### ğŸ“‹ è¯¦ç»†é¢„è®¢è®°å½•")
                    
                    # é€‰æ‹©è¦æ˜¾ç¤ºçš„åˆ—
                    display_columns = ['æ—¥æœŸ', 'æ¡Œç‰Œå·', 'åŒ¹é…çŠ¶æ€', 'åŒ¹é…ç±»å‹']
                    if 'æ”¯ä»˜åˆè®¡' in customer_data.columns:
                        display_columns.append('æ”¯ä»˜åˆè®¡')
                    if 'ä¸‹å•æ—¶é—´' in customer_data.columns:
                        display_columns.append('ä¸‹å•æ—¶é—´')
                    
                    available_columns = [col for col in display_columns if col in customer_data.columns]
                    
                    if available_columns:
                        display_data = customer_data[available_columns].copy()
                        
                        # æŒ‰æ—¥æœŸæ’åº
                        if 'æ—¥æœŸ' in display_data.columns:
                            display_data = display_data.sort_values('æ—¥æœŸ', ascending=False)
                        
                        st.dataframe(
                            display_data,
                            use_container_width=True,
                            hide_index=True
                        )
                        
                        # å¯¼å‡ºè¯¥å®¢æˆ·çš„æ•°æ®
                        if st.button(f"ğŸ“¥ å¯¼å‡º {customer_name} çš„æ•°æ®", use_container_width=True):
                            # åˆ›å»ºExcelæ–‡ä»¶
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                display_data.to_excel(writer, sheet_name=f'{customer_name}_é¢„è®¢è®°å½•', index=False)
                            
                            excel_data = output.getvalue()
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            filename = f"{customer_name}_é¢„è®¢åˆ†æ_{timestamp}.xlsx"
                            
                            st.download_button(
                                label=f"ä¸‹è½½ {customer_name} çš„é¢„è®¢æ•°æ®",
                                data=excel_data,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.warning("æ— å¯æ˜¾ç¤ºçš„è¯¦ç»†æ•°æ®")
            else:
                st.info("è¯·åœ¨å·¦ä¾§é€‰æ‹©ä¸€ä¸ªé¢„è®¢äººè¿›è¡Œåˆ†æ")
                
                # æ˜¾ç¤ºæ•´ä½“ç»Ÿè®¡æ¦‚è§ˆ
                st.markdown("### ğŸ“Š æ•´ä½“æ•°æ®æ¦‚è§ˆ")
                
                # æœ€æ´»è·ƒçš„é¢„è®¢äººTop 10
                if 'é¢„è®¢äºº' in self.merged_df.columns:
                    # ä½¿ç”¨æ ‡å‡†åŒ–åçš„å§“åè¿›è¡Œç»Ÿè®¡
                    standardized_names = self.merged_df['é¢„è®¢äºº'].apply(self.normalize_customer_name)
                    valid_customers = standardized_names.dropna()
                    top_customers = valid_customers.value_counts().head(10)
                    
                    if not top_customers.empty:
                         st.markdown("#### ğŸ† æœ€æ´»è·ƒé¢„è®¢äºº (Top 10)")
                         
                         fig_top = px.bar(
                             x=top_customers.values,
                             y=top_customers.index,
                             orientation='h',
                             title="æœ€æ´»è·ƒçš„é¢„è®¢äººæ’è¡Œæ¦œ",
                             labels={'x': 'é¢„è®¢æ¬¡æ•°', 'y': 'é¢„è®¢äºº'}
                         )
                         fig_top.update_layout(
                             height=400,
                             font=dict(family="Microsoft YaHei, SimHei, sans-serif")
                         )
                         st.plotly_chart(fig_top, use_container_width=True)

def main():
    # åœ¨åµŒå…¥ä¸»åº”ç”¨æ—¶é¿å…é‡å¤è®¾ç½®é¡µé¢é…ç½®
    try:
        st.set_page_config(
            page_title="é¹­åºœé¢„å®šåŒ¹é…å·¥å…· v2.0",
            page_icon="ğŸ“Š",
            layout="wide",
            initial_sidebar_state="collapsed"
        )
    except Exception:
        # å·²åœ¨ä¸»åº”ç”¨ä¸­è®¾ç½®è¿‡ï¼Œå¿½ç•¥
        pass
    
    # ç®€æ´ç°ä»£çš„CSSæ ·å¼
    st.markdown("""
    <style>
    .main {
        padding: 1rem 2rem;
        max-width: 1200px;
        margin: 0 auto;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 0;
        border-bottom: 1px solid #e1e5e9;
        background: transparent;
    }
    .stTabs [data-baseweb="tab"] {
        height: 3rem;
        padding: 0 2rem;
        background: transparent;
        border: none;
        border-bottom: 2px solid transparent;
        font-weight: 500;
        color: #64748b;
        font-size: 14px;
    }
    .stTabs [aria-selected="true"] {
        background: transparent;
        color: #0f172a;
        border-bottom-color: #3b82f6;
    }
    .stFileUploader {
        border: 2px dashed #cbd5e1;
        border-radius: 8px;
        padding: 2rem;
        background: #f8fafc;
        transition: all 0.2s ease;
    }
    .stFileUploader:hover {
        border-color: #3b82f6;
        background: #f1f5f9;
    }
    .metric-container {
        background: white;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 1.5rem;
        margin: 0.5rem 0;
    }
    .stAlert {
        border-radius: 8px;
        border: none;
    }
    .stAlert[data-baseweb="notification"][kind="success"] {
        background: #ecfdf5;
        color: #059669;
        border-left: 4px solid #059669;
    }
    .stAlert[data-baseweb="notification"][kind="error"] {
        background: #fef2f2;
        color: #dc2626;
        border-left: 4px solid #dc2626;
    }
    .stAlert[data-baseweb="notification"][kind="warning"] {
        background: #fffbeb;
        color: #d97706;
        border-left: 4px solid #d97706;
    }
    .stAlert[data-baseweb="notification"][kind="info"] {
        background: #eff6ff;
        color: #2563eb;
        border-left: 4px solid #2563eb;
    }
    .stDataFrame {
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        overflow: hidden;
    }
    h1, h2, h3 {
        color: #0f172a;
        font-weight: 600;
    }
    .stButton > button {
        border-radius: 6px;
        font-weight: 500;
        transition: all 0.2s ease;
    }
    .stButton > button[kind="primary"] {
        background: #3b82f6;
        border: none;
    }
    .stButton > button[kind="primary"]:hover {
        background: #2563eb;
        transform: translateY(-1px);
    }
    </style>
    """, unsafe_allow_html=True)
    
    # ç®€æ´çš„æ ‡é¢˜
    st.markdown("""
    <div style='text-align: center; margin-bottom: 2rem;'>
        <h1 style='color: #0f172a; font-weight: 600; font-size: 2rem; margin: 0;'>é¹­åºœé¢„å®šåŒ¹é…å·¥å…·</h1>
    </div>
    """, unsafe_allow_html=True)
    
    # åˆå§‹åŒ–åº”ç”¨
    if 'app' not in st.session_state:
        st.session_state.app = ReservationMatcherWeb()
    
    app = st.session_state.app
    
    # ä¸‰ä¸ªä¸»è¦æ ‡ç­¾é¡µ
    tab1, tab2, tab3 = st.tabs(["ğŸ“ æ–‡ä»¶å¤„ç†", "ğŸ“Š ç»“æœæŸ¥çœ‹", "ğŸ“ˆ æ•°æ®åˆ†æ"])
    
    # æ ‡ç­¾é¡µå†…å®¹
    with tab1:
        # æ–‡ä»¶ä¸Šä¼ å’Œæ•°æ®åŒ¹é…åˆå¹¶
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("ğŸ“¤ æ–‡ä»¶ä¸Šä¼ ")
            app.load_files()
            
        with col2:
            st.subheader("âš¡ æ•°æ®åŒ¹é…")
            
            # éªŒè¯æ–‡ä»¶
            is_valid, message = app.validate_files()
            
            if not is_valid:
                st.warning(message)
            else:
                st.success("æ–‡ä»¶å·²å°±ç»ª")
                
                if st.button("ğŸš€ å¼€å§‹åŒ¹é…", type="primary", use_container_width=True):
                    with st.spinner("åŒ¹é…ä¸­..."):
                        success, result_message = app.match_data()
                        
                    if success:
                        st.success(result_message)
                        st.info("è¯·åˆ‡æ¢åˆ°'ç»“æœæŸ¥çœ‹'æ ‡ç­¾é¡µ")
                    else:
                        st.error(result_message)
    
    with tab2:
        # æŸ¥çœ‹ç»“æœå’Œå¯¼å‡ºåˆå¹¶
        col1, col2 = st.columns([3, 1])
        
        with col1:
            app.display_results()
            
        with col2:
            st.subheader("ğŸ“¥ å¯¼å‡º")
            app.export_results()
    
    with tab3:
        # æ•°æ®åˆ†ææ ‡ç­¾é¡µ
        app.show_data_analysis()
    


if __name__ == "__main__":
    main()